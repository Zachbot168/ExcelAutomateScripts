import openpyxl

def remove_rows_below_departure_arrival(file_path, output_path):
    among_us = openpyxl.load_workbook(file_path)

    for sheet in among_us.worksheets:
        # Find the row containing the titles "Departure" and "Arrival"
        target_row = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ["DEPARTURE", "ARRIVAL"]:
                    target_row = cell.row
                    break
            if target_row:
                break

        # If "Departure" or "Arrival" was found, remove rows below it
        if target_row:
            max_row = sheet.max_row
            for row in range(target_row, max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).value = None

    among_us.save(output_path)
    print(f"Removed rows below 'Departure' and 'Arrival', and saved the new file to {output_path}")

input_file = input("Please enter your input file path: ")
output_file = input("Please enter your output file path: ")
remove_rows_below_departure_arrival(input_file, output_file)