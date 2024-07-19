import openpyxl

def unmerge_cells(file_path, output_path):
    among_us = openpyxl.load_workbook(file_path)

    for sheet in among_us.worksheets:
        for merged_cell in list(sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(merged_cell.coord)

            merged_value = []
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value:
                        merged_value.append(str(cell_value))
            merged_value = " ".join(merged_value)

            sheet.unmerge_cells(merged_cell.coord)

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = merged_value

    among_us.save(output_path)
    print(f"Unmerged cells and saved the new file to " + output_file)


input_file = input("please put your input file path ")
output_file = input("please put your output file path ")
unmerge_cells(input_file, output_file)
