import openpyxl
from openpyxl.utils import get_column_letter


def remove_tz_from_time_column(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.worksheets:
        # Find the column index for "time"
        time_col = None
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value == "time":
                time_col = col_idx
                break

        if time_col is None:
            print(f"'time' column not found in sheet {sheet.title}")
            continue

        # Iterate through the rows and update the "time" column
        for row in sheet.iter_rows(min_row=2, min_col=time_col, max_col=time_col):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "T" in cell.value and "Z" in cell.value:
                    cell.value = cell.value.replace("T", " ").replace("Z", "")

    # Save the modified workbook
    wb.save(output_path)
    print(f"Updated 'time' column and saved to {output_path}")


input_file = input("Please enter your input file path: ")
output_file = input("Please enter your output file path: ")
remove_tz_from_time_column(input_file, output_file)
