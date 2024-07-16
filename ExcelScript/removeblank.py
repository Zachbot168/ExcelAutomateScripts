import openpyxl


def remove_rows_with_empty_columns(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows_to_remove = []

        # Check if any of the first three columns are empty for each row
        for row_idx in range(1, sheet.max_row + 1):
            is_empty = any(sheet.cell(row=row_idx, column=col_idx).value is None for col_idx in range(1, 4))
            if is_empty:
                rows_to_remove.append(row_idx)

        # Remove identified rows in reverse order to avoid shifting issues
        for row_idx in reversed(rows_to_remove):
            sheet.delete_rows(row_idx)

    # Save the modified workbook
    wb.save(output_path)
    print(f"Removed rows with any empty first three columns and saved to {output_path}")


input_file = input("Please enter your input file path: ")
output_file = input("Please enter your output file path: ")
remove_rows_with_empty_columns(input_file, output_file)
