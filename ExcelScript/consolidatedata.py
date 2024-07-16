import openpyxl


def consolidate_sheets(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)

    # Create a new workbook for consolidated data
    consolidated_wb = openpyxl.Workbook()
    consolidated_sheet = consolidated_wb.active
    consolidated_sheet.title = 'Consolidated Data'

    header_written = False  # Flag to track if header has been written

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Skip sheets without data
        if sheet.max_row == 0:
            continue

        # Copy data from each sheet to consolidated sheet
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not header_written:
                # Write header from the first sheet only
                consolidated_sheet.append(row)
                header_written = True
            elif row_idx > 1:  # Skip the header row for subsequent sheets
                consolidated_sheet.append(row)

    # Save the consolidated workbook
    consolidated_wb.save(output_path)
    print(f"Consolidated all sheets into one and saved to {output_path}")

input_file = input("Please enter your input file path: ")
output_file = input("Please enter your output file path: ")
consolidate_sheets(input_file, output_file)
