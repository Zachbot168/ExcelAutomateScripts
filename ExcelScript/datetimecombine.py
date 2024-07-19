import openpyxl
from datetime import datetime, time

def reformat_date(date_str):
    try:
        # Try parsing common date formats
        return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        try:
            return datetime.strptime(date_str, "%m/%d/%Y %I:%M:%S %p")
        except ValueError:
            return None

def update_date_with_time(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)

    for sheet in wb.worksheets:
        # Find the column indices for "DATE" and "TIME START"
        date_col = None
        time_start_col = None
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value == "DATE":
                date_col = col_idx
            elif cell_value == "TIME STARTED (UTC)":
                time_start_col = col_idx

        if date_col is None or time_start_col is None:
            print(f"'DATE' or 'TIME START' column not found in sheet {sheet.title}")
            continue

        rows_to_delete = []

        # Update the "DATE" column to include only the date from "DATE" and time from "TIME START"
        for row_idx in range(2, sheet.max_row + 1):
            date_value = sheet.cell(row=row_idx, column=date_col).value
            time_start_value = sheet.cell(row=row_idx, column=time_start_col).value

            if time_start_value is None:
                rows_to_delete.append(row_idx)
                continue

            if isinstance(date_value, str):
                date_value = reformat_date(date_value)

            if date_value is None:
                rows_to_delete.append(row_idx)
                continue

            try:
                if isinstance(date_value, datetime) and isinstance(time_start_value, time):
                    new_datetime = datetime.combine(date_value.date(), time_start_value)
                    sheet.cell(row=row_idx, column=date_col).value = new_datetime
                elif isinstance(date_value, datetime) and isinstance(time_start_value, str):
                    time_start_value = datetime.strptime(time_start_value, "%H:%M:%S").time()
                    new_datetime = datetime.combine(date_value.date(), time_start_value)
                    sheet.cell(row=row_idx, column=date_col).value = new_datetime
            except ValueError as e:
                print(f"Error parsing date/time in row {row_idx}: {e}")
                rows_to_delete.append(row_idx)

        # Delete rows with empty "TIME START" in reverse order to avoid shifting issues
        for row_idx in reversed(rows_to_delete):
            sheet.delete_rows(row_idx)

    # Save the modified workbook
    wb.save(output_path)
    print(f"Updated 'DATE' column with 'TIME START' time and saved to {output_path}")

input_file = input("Please enter your input file path: ")
output_file = input("Please enter your output file path: ")
update_date_with_time(input_file, output_file)
