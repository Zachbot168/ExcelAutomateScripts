import openpyxl

def compare_excel_sheets(file1, file2, output_file):
    # Load workbooks and sheets
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    ws1 = wb1.active
    ws2 = wb2.active

    # Find headers in both sheets
    headers1 = {cell.value: idx for idx, cell in enumerate(ws1[1])}
    headers2 = {cell.value: idx for idx, cell in enumerate(ws2[1])}

    # Ensure necessary columns exist in both sheets
    if 'STATION' not in headers1 or 'DATE' not in headers1 or 'airport_code' not in headers2 or 'time' not in headers2:
        raise ValueError('One of the required columns is missing in one of the files.')

    # Add headers of the second sheet to the first sheet
    for col_idx, header in enumerate(ws2[1], start=len(headers1) + 1):
        ws1.cell(row=1, column=col_idx, value=header.value)

    # Create a map of the second sheet's rows for fast lookup
    rows2 = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        airport_code = row[headers2['airport_code']]
        time = row[headers2['time']]
        if (airport_code, time) not in rows2:
            rows2[(airport_code, time)] = []
        rows2[(airport_code, time)].append(row)

    # Iterate through the first sheet and find matching rows
    for row in ws1.iter_rows(min_row=2):
        station = row[headers1['STATION']].value
        date = row[headers1['DATE']].value

        if (station, date) in rows2:
            for match_row in rows2[(station, date)]:
                # Append matched row from second sheet to first sheet
                for col_idx, value in enumerate(match_row, start=len(headers1) + 1):
                    ws1.cell(row=row[0].row, column=col_idx, value=value)

    # Save the updated workbook
    wb1.save(output_file)

# Example usage
compare_excel_sheets('/Users/zacharylee/Downloads/Newly Corrected Final Consolidated Ligthning Alert Monitoring July 2023 to July 2024.xlsx', '/Users/zacharylee/Downloads/new_syntaxcorrect_new_weather_observations_flightaware.xlsx', '/Users/zacharylee/Downloads/Ultimate Newly Corrected Final Consolidated Ligthning Alert Monitoring July 2023 to July 2024.xlsx')
