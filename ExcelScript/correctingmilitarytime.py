import openpyxl
from datetime import datetime

def convert_military_to_12_hour(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column):
        time_cell = row[1]  # Assuming "time" column is the second column (index 1)

        if time_cell.value:
            try:
                # Convert the string to a datetime object
                time_value = datetime.strptime(time_cell.value, "%Y-%m-%d %H:%M:%S")

                # Format it to 12-hour time with AM/PM
                formatted_time = time_value.strftime("%Y-%m-%d %I:%M:%S %p")

                # Update the cell with the new format
                time_cell.value = formatted_time

            except ValueError:
                print(f"Skipping row {row[0].row} due to format error: {time_cell.value}")

    wb.save(output_path)
    print(f"Converted times to 12-hour format and saved to {output_path}")

# User input
file_path = input("Please enter the path to the Excel file: ")
output_path = input("Please enter the path to save the output Excel file: ")

convert_military_to_12_hour(file_path, output_path)
