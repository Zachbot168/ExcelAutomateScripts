import openpyxl
from datetime import datetime

def reformat_date_to_custom_format(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)

    for sheet in wb.worksheets:
        # Find the column index for "DATE"
        date_col = None
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value == "DATE":
                date_col = col_idx
                break

        if date_col is None:
            print(f"'DATE' column not found in sheet {sheet.title}")
            continue

        rows_to_delete = []

        # Update the "DATE" column to the custom format
        for row_idx in range(2, sheet.max_row + 1):
            date_value = sheet.cell(row=row_idx, column=date_col).value

            if date_value is None:
                continue

            try:
                if isinstance(date_value, datetime):
                    # Format datetime object to custom format
                    new_format = date_value.strftime('%Y-%m-%dT%H:%M:%SZ')
                    sheet.cell(row=row_idx, column=date_col).value = new_format
                elif isinstance(date_value, str):
                    # Parse different date formats if needed
                    try:
                        date_value = datetime.strptime(date_value, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        date_value = datetime.strptime(date_value, "%m/%d/%Y %I:%M:%S %p")

                    new_format = date_value.strftime('%Y-%m-%dT%H:%M:%SZ')
                    sheet.cell(row=row_idx, column=date_col).value = new_format
                else:
                    continue

            except ValueError as e:
                print(f"Error parsing date in row {row_idx}: {date_value}, {e}")
                rows_to_delete.append(row_idx)

        # Remove identified rows in reverse order to avoid shifting issues
        for row_idx in reversed(rows_to_delete):
            sheet.delete_rows(row_idx)

    # Save the modified workbook
    wb.save(output_path)
    print(f"Reformatted 'DATE' column to custom format and saved to {output_path}")

input_file = input("Please enter your input file path: ")
output_file = input("Please enter your output file path: ")
reformat_date_to_custom_format(input_file, output_file)
