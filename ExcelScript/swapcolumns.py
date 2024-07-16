import openpyxl

def adjust_date_station_columns(file_path, output_path):
    among_us = openpyxl.load_workbook(file_path)

    for sheet in among_us.worksheets:
        date_col = None
        station_col = None

        # Find column indices for "DATE" and "STATION"
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value == "DATE":
                date_col = col_idx
            elif cell_value == "STATION":
                station_col = col_idx

        # Swap columns if necessary
        if date_col and station_col:
            if date_col != 1:
                # Store values in DATE column
                date_values = []
                for row_idx in range(1, sheet.max_row + 1):
                    date_values.append(sheet.cell(row=row_idx, column=date_col).value)
                # Move STATION to DATE column
                for row_idx in range(1, sheet.max_row + 1):
                    station_value = sheet.cell(row=row_idx, column=station_col).value
                    sheet.cell(row=row_idx, column=date_col).value = station_value
                # Move DATE values to STATION column
                for row_idx in range(1, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=station_col).value = date_values[row_idx - 1]

    among_us.save(output_path)
    print(f"Adjusted DATE and STATION columns as needed, and saved the new file to {output_path}")

input_file = input("Please enter your input file path: ")
output_file = input("Please enter your output file path: ")
adjust_date_station_columns(input_file, output_file)
